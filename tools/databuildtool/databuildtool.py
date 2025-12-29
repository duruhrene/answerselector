# ------------------------------------------------------------------------------
# AnswerSelector Data Build Tool v1.0.0 (2025-12-30)
# Copyright (c) 2025-2026 Duruhrene. All rights reserved.
# ------------------------------------------------------------------------------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sys
import threading
import json
import sqlite3
import numpy as np
from pathlib import Path
from typing import Optional, List, Any, Dict
import openpyxl

# Additional imports for EmbedManager
try:
    import onnxruntime as ort
    from tokenizers import Tokenizer
except ImportError:
    ort = None
    Tokenizer = None

# --- EmbedManager Class Integration ---
class EmbedManager:
    """
    Manages ONNX-based embedding models for semantic retrieval.
    Handles tokenization, inference, pooling, and L2 normalization.
    """

    def __init__(self, model_dir: Optional[Any] = None) -> None:
        self.model_dir = Path(model_dir) if model_dir else self._default_model_dir()
        
        self.tokenizer: Optional[Any] = None
        self.session: Optional[Any] = None
        
        # Model Configuration
        self.hidden_size = 768
        self.max_length = 512
        self.output_name = "last_hidden_state"
        self.use_pooling = True
        self.top_k = 15

    def _default_model_dir(self) -> Path:
        base_dir = Path(__file__).resolve().parent
        return base_dir / "model"

    def load_model(self) -> bool:
        if ort is None or Tokenizer is None:
            print("Error: onnxruntime or tokenizers not installed.")
            return False

        model_path = self.model_dir / "model.onnx"
        tok_path = self.model_dir / "tokenizer.json"

        if not model_path.exists() or not tok_path.exists():
            print(f"Error: Model files not found in {self.model_dir}")
            return False

        try:
            self.tokenizer = Tokenizer.from_file(str(tok_path))
            self.tokenizer.enable_truncation(max_length=self.max_length)
            self.tokenizer.enable_padding(length=self.max_length)

            providers = ['CPUExecutionProvider']
            self.session = ort.InferenceSession(str(model_path), providers=providers)
            
            return True
        except Exception as e:
            print(f"Error loading ONNX model: {e}")
            return False

    def unload_model(self) -> None:
        self.tokenizer = None
        self.session = None

    def _mean_pooling(self, last_hidden_state: np.ndarray, attention_mask: np.ndarray) -> np.ndarray:
        mask_expanded = np.expand_dims(attention_mask, axis=-1).astype(float)
        sum_embeddings = np.sum(last_hidden_state * mask_expanded, axis=1)
        sum_mask = np.clip(np.sum(mask_expanded, axis=1), a_min=1e-9, a_max=None)
        return sum_embeddings / sum_mask

    def get_embedding(self, text: str) -> Optional[np.ndarray]:
        if not text:
            return None

        if self.session is None or self.tokenizer is None:
            if not self.load_model():
                return None

        try:
            encoded = self.tokenizer.encode(text)
            input_ids = np.array([encoded.ids], dtype=np.int64)
            attention_mask = np.array([encoded.attention_mask], dtype=np.int64)

            inputs = {
                "input_ids": input_ids,
                "attention_mask": attention_mask
            }
            outputs = self.session.run([self.output_name], inputs)
            raw_output = outputs[0]

            if self.use_pooling:
                sentence_embedding = self._mean_pooling(raw_output, attention_mask)[0]
            else:
                sentence_embedding = raw_output[0]

            norm = np.linalg.norm(sentence_embedding)
            if norm > 0:
                sentence_embedding = sentence_embedding / norm
                
            return sentence_embedding
        except Exception as e:
            print(f"Embedding error: {e}")
            return None

class DataBuilderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Build Tool for AnswerSelector v1")
        self.root.geometry("600x500")
        
        self.InitUI()
        self._set_defaults()

    def InitUI(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. Excel File
        ttk.Label(main_frame, text="엑셀 파일:").pack(anchor=tk.W, pady=(5, 0))
        
        hbox1 = ttk.Frame(main_frame)
        hbox1.pack(fill=tk.X, pady=5)
        self.excel_path_var = tk.StringVar()
        ttk.Entry(hbox1, textvariable=self.excel_path_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(hbox1, text="찾기", command=self.browse_excel).pack(side=tk.RIGHT)

        # 2. Model Directory
        ttk.Label(main_frame, text="모델 폴더:").pack(anchor=tk.W, pady=(10, 0))
        
        hbox2 = ttk.Frame(main_frame)
        hbox2.pack(fill=tk.X, pady=5)
        self.model_dir_var = tk.StringVar()
        ttk.Entry(hbox2, textvariable=self.model_dir_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(hbox2, text="찾기", command=self.browse_model).pack(side=tk.RIGHT)

        # 3. Output Directory
        ttk.Label(main_frame, text="출력(DB) 폴더:").pack(anchor=tk.W, pady=(10, 0))
        
        hbox3 = ttk.Frame(main_frame)
        hbox3.pack(fill=tk.X, pady=5)
        self.output_dir_var = tk.StringVar()
        ttk.Entry(hbox3, textvariable=self.output_dir_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(hbox3, text="찾기", command=self.browse_output).pack(side=tk.RIGHT)

        # 4. Build Button
        self.btn_build = ttk.Button(main_frame, text="데이터 빌드", command=self.on_build)
        self.btn_build.pack(pady=20)

        # 5. Log Area
        self.log_text = tk.Text(main_frame, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _set_defaults(self):
        # Handle frozen state (EXE vs Script)
        if getattr(sys, 'frozen', False):
            app_dir = Path(sys.executable).parent
        else:
            app_dir = Path(__file__).resolve().parent

        self.excel_path_var.set(str(app_dir))
        self.model_dir_var.set(str(app_dir))
        self.output_dir_var.set(str(app_dir))

    def browse_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.excel_path_var.set(filename)

    def browse_model(self):
        dirname = filedialog.askdirectory()
        if dirname:
            self.model_dir_var.set(dirname)

    def browse_output(self):
        dirname = filedialog.askdirectory()
        if dirname:
            self.output_dir_var.set(dirname)

    def log(self, msg):
        def _log():
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
        self.root.after(0, _log)

    def on_build(self):
        excel_path = self.excel_path_var.get()
        model_dir = self.model_dir_var.get()
        output_dir = self.output_dir_var.get()

        if not excel_path or not model_dir or not output_dir:
            messagebox.showerror("Error", "모든 경로를 선택해주세요.")
            return

        self.btn_build.config(state=tk.DISABLED)
        self.log("빌드 시작...")

        thread = threading.Thread(target=self.run_build, args=(excel_path, model_dir, output_dir))
        thread.start()

    def run_build(self, excel_path, model_dir, output_dir):
        try:
            self.log(f"엑셀 읽기: {excel_path}")
            
            # --- Load Workbook ---
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
            except Exception as e:
                self.log(f"엑셀 열기 실패: {e}")
                self.root.after(0, lambda: self.btn_build.config(state=tk.NORMAL))
                return

            # --- Initialize AI ---
            if not EmbedManager:
                self.log("오류: EmbedManager(onnxruntime/tokenizers) 로드 실패.")
                self.root.after(0, lambda: self.btn_build.config(state=tk.NORMAL))
                return

            self.log("AI 모델 로딩 중...")
            embed_manager = EmbedManager(model_dir=model_dir)
            if not embed_manager.load_model():
                self.log("경고: AI 모델 로드 실패. 임베딩은 비어있게 됩니다.")

            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)

            self._process_answer_embed(wb, output_path, embed_manager)
            self._process_agencies(wb, output_path)
            self._process_intro_closing(wb, output_path)
            self._process_conjunctions(wb, output_path)

            self.log("\n[완료] 모든 데이터 생성이 끝났습니다.")
            self.root.after(0, lambda: messagebox.showinfo("성공", "데이터 빌드 완료!"))

        except Exception as e:
            self.log(f"\n[오류] {e}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.root.after(0, lambda: self.btn_build.config(state=tk.NORMAL))

    def _get_sheet_data(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = rows[0]
        data = []
        for r in rows[1:]:
            # Map headers to values
            row_dict = {}
            for i, h in enumerate(headers):
                if h:
                    row_dict[h] = r[i] if i < len(r) else None
            data.append(row_dict)
        return data

    def _process_answer_embed(self, wb, db_dir, embed_manager):
        self.log("\nProcessing 'answerembed' sheet...")
        data = self._get_sheet_data(wb, "answerembed")
        if data is None:
            self.log("시트 없음: answerembed")
            return

        db_path = db_dir / "answersembed.db"
        if db_path.exists():
            db_path.unlink()

        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE answerembed (
                id INTEGER PRIMARY KEY,
                code TEXT,
                cat1 TEXT,
                cat2 TEXT,
                cat3 TEXT,
                title TEXT,
                maintext TEXT,
                agency1 TEXT,
                agency2 TEXT,
                embedding TEXT
            )
        """)
        
        count = 0
        total = len(data)
        
        for row in data:
            emb_json = "[]"
            maintext = row.get('maintext')
            if embed_manager and embed_manager.session:
                text_to_embed = str(maintext) if maintext else ""
                emb_vector = embed_manager.get_embedding(text_to_embed)
                if emb_vector is not None:
                    emb_json = json.dumps(emb_vector.tolist())
            
            conn.execute("""
                INSERT INTO answerembed (id, code, cat1, cat2, cat3, title, maintext, agency1, agency2, embedding)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row.get('id'), row.get('code'), row.get('cat1'), row.get('cat2'), row.get('cat3'),
                row.get('title'), row.get('maintext'), row.get('agency1'), row.get('agency2'),
                emb_json
            ))
            count += 1
            if count % 10 == 0:
                self.log(f"Processed {count}/{total}...")

        conn.commit()
        conn.close()
        self.log(f"Saved {count} records to answersembed.db")

    def _process_agencies(self, wb, db_dir):
        self.log("\nProcessing 'agencies' sheet...")
        data = self._get_sheet_data(wb, "agencies")
        if data is None: return

        db_path = db_dir / "agencies.db"
        if db_path.exists(): db_path.unlink()

        conn = sqlite3.connect(db_path)
        conn.execute("CREATE TABLE agencies (id INTEGER PRIMARY KEY, name TEXT, website TEXT, tel TEXT, paid TEXT)")
        
        count = 0
        for row in data:
            conn.execute("INSERT INTO agencies VALUES (?, ?, ?, ?, ?)",
                         (row.get('id'), row.get('name'), row.get('website'), row.get('tel'), row.get('paid')))
            count += 1
        conn.commit()
        conn.close()
        self.log(f"Saved {count} agencies.")

    def _process_intro_closing(self, wb, db_dir):
        self.log("\nProcessing 'introclosing' sheet...")
        data = self._get_sheet_data(wb, "introclosing")
        if data is None: return

        db_path = db_dir / "introclosing.db"
        if db_path.exists(): db_path.unlink()

        conn = sqlite3.connect(db_path)
        conn.execute("CREATE TABLE introclosing (id INTEGER PRIMARY KEY, type TEXT, cat TEXT, text TEXT)")
        
        count = 0
        for row in data:
            conn.execute("INSERT INTO introclosing VALUES (?, ?, ?, ?)",
                         (row.get('id'), row.get('type'), row.get('cat'), row.get('text')))
            count += 1
        conn.commit()
        conn.close()
        self.log(f"Saved {count} intro/closing.")

    def _process_conjunctions(self, wb, db_dir):
        self.log("\nProcessing 'conjunctions' sheet...")
        data = self._get_sheet_data(wb, "conjunctions")
        if data is None: return

        json_path = db_dir / "conjunctions.json"
        
        # 'conj' column extraction
        conjs = []
        for row in data:
            val = row.get('conj')
            if val:
                conjs.append(str(val))
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(conjs, f, ensure_ascii=False, indent=4)
        self.log(f"Saved {len(conjs)} conjunctions.")

if __name__ == "__main__":
    root = tk.Tk()
    app = DataBuilderApp(root)
    root.mainloop()
